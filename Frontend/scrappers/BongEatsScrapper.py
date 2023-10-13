#Bong Eats

from bs4 import BeautifulSoup
import requests
import re
from openpyxl import load_workbook
import openpyxl

link = str(input('Enter the link: '))
type = str(input("lunch,dinner(1)/ snacks(2)/ snacks,lunch,dinner(3)/ dessert(4)? : "))
diet = str(input('vegeterian(1)/non-veg(2)? : '))

if(type == '1'):
    type='lunch,dinner'
elif(type == '2'):
    type='snacks'
elif(type == '3'):
    type='snacks,lunch,dinner'
elif(type == '4'):
    type='dessert'

if(diet == '1'):
    diet='vegeterian'
elif(diet == '2'):
    diet='non-veg'


#starting the scrapping
html_texts = requests.get(link).text

soup = BeautifulSoup(html_texts, 'lxml')

#to find the name of the recipe
recipe = soup.find('div', class_='text-container').h1
recipe = recipe.text

#to find the servings
serving = soup.find('div', class_='column-3 w-col w-col-6')
serving = serving.p.text[0]

#to find the cooking time
cook_time = soup.find('div', class_='column-4 w-col w-col-6')
cook_time = cook_time.p.text

cook_time = re.split('\s', cook_time)
time = float(cook_time[0])

##to convert the time into minute
if(cook_time[1] == 'hours' or cook_time[1] == 'hour'):
    time *= 60

#to find the ingredients
ingredients = soup.findAll('ul')

for ing in ingredients:
    print(ing)

ingredients.pop()
ingredients.pop()
final_ingredients = ''

for ing in ingredients:
    for li in ing.find_all("li"):

        if(li.strong and li.strong.text[-1] != 'ml'):
            # for quantity
            quantity = li.strong.text
            quantity = re.split('\s', quantity)

            f = len(quantity)

            #print(quantity)
            #print(len(quantity))

            if (len(quantity) > 1):
                if (quantity[1] == 'kg'):
                    if (quantity[0] == 'Â½' or quantity[0] == '1â2'):
                        quantity[0] = 0.5
                    elif (quantity[0] == 'Â¼' or quantity[0] == '1â4'):
                        quantity[0] = 0.25
                    elif (quantity[0] == 'Â¾'):
                        quantity[0] = 0.75
                    elif (quantity[0] == '1Â½'):
                        quantity[0] = 1.5
                    quantity[0] = str(float(quantity[0]) * 1000) + 'g'

                elif (quantity[1] == 'tsp'):
                    if (quantity[0] == 'Â½' or quantity[0] == '1â2'):
                        quantity[0] = 0.5
                    elif (quantity[0] == 'Â¼' or quantity[0] == '1â4'):
                        quantity[0] = 0.25
                    elif(quantity[0] == 'Â¾'):
                        quantity[0] = 0.75
                    elif (quantity[0] == '1Â½'):
                        quantity[0] = 1.5
                    quantity[0] = str(float(quantity[0]) * 4.2) + 'g'

                elif (quantity[1] == 'tbsp'):
                    if (quantity[0] == 'Â½' or quantity[0] == '1â2'):
                        quantity[0] = 0.5
                    elif (quantity[0] == 'Â¼' or quantity[0] == '1â4'):
                        quantity[0] = 0.25
                    elif (quantity[0] == 'Â¾'):
                        quantity[0] = 0.75
                    elif (quantity[0] == '1Â½'):
                        quantity[0] = 1.5
                    quantity[0] = str(float(quantity[0]) * 15) + 'g'

                elif (quantity[1] == 'cup'):
                    if (quantity[0] == 'Â½' or quantity[0] == '1â2'):
                        quantity[0] = 0.5
                    elif (quantity[0] == 'Â¼' or quantity[0] == '1â4'):
                        quantity[0] = 0.25
                    elif (quantity[0] == 'Â¾'):
                        quantity[0] = 0.75
                    elif (quantity[0] == '1Â½'):
                        quantity[0] = 1.5
                    quantity[0] = str(float(quantity[0]) * 250) + 'g'

                elif (quantity[1] == 'pinch'):
                    if (quantity[0] == 'Â½' or quantity[0] == '1â2'):
                        quantity[0] = 0.5
                    elif (quantity[0] == 'Â¼' or quantity[0] == '1â4'):
                        quantity[0] = 0.25
                    elif (quantity[0] == 'Â¾'):
                        quantity[0] = 0.75
                    elif (quantity[0] == '1Â½'):
                        quantity[0] = 1.5
                    quantity[0] = str(float(quantity[0]) * 1.5) + 'g'

                else:
                    if (quantity[0] == 'Â½' or quantity[0] == '1â2'):
                        quantity[0] = 0.5
                    elif (quantity[0] == 'Â¼' or quantity[0] == '1â4'):
                        quantity[0] = 0.25
                    elif (quantity[0] == 'Â¾'):
                        quantity[0] = 0.75
                    elif (quantity[0] == '1Â½'):
                        quantity[0] = 1.5
                    quantity[0] = str(quantity[0]) + 'g'

            else:
                if (quantity[0] == 'Â½' or quantity[0] == '1â2'):
                    quantity[0] = 0.5
                elif (quantity[0] == 'Â¼' or quantity[0] == '1â4'):
                    quantity[0] = 0.25
                elif (quantity[0] == 'Â¾'):
                    quantity[0] = 0.75
                elif (quantity[0] == '1Â½'):
                    quantity[0] = 1.5
                quantity[0] = quantity[0] + 'g'

            # for ingredients name
            names = re.split('\s', li.text)

            n = quantity[0]
            b = False
            k = 0

            for name in names:

                f -= 1

                if (f == 0): b = True
                if (name[0] == '('): b = False
                if (b):
                    n += ' ' + name
            final_ingredients += n + ', '

final_ingredients = final_ingredients[:len(final_ingredients)-2]

print(f'Name: {recipe}\nIngredients: {final_ingredients}\nTime(s): {time}\nDiet: {diet}\nServing: {serving}\nType: {type}\nLink: {link}\n')

sa = str(input('Save? y/n : '))

if(sa == 'y'):
    # save data in a excel file
    #wb = openpyxl.Workbook()
    wb = load_workbook('Recipe.xlsx')
    ws = wb.active

    data = (recipe, final_ingredients, time, diet, serving, type, link)

    ws.append(data)
    wb.save('Recipe.xlsx')
