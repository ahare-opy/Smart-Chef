import sys
import traceback

import pymongo
import openpyxl

wb = openpyxl.load_workbook('Final_Recipe_Dataset.xlsx')

ws = wb.active

temparray = []
for row in range(0, ws.max_row):
    values = []
    for col in ws.iter_cols(1, ws.max_column):
        values.append(col[row].value)
    temparray.append(values)
#321
#168
print(temparray[0])




