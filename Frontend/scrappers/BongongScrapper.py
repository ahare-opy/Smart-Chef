#Bong Gong

from bs4 import BeautifulSoup
import requests
import re
from openpyxl import load_workbook
import openpyxl
from translate import Translator

link = str(input('Enter the link: '))
type = str(input("lunch,dinner(1)/ snacks(2)/ snacks,lunch,dinner(3)/ dessert(4)? : "))
diet = str(input('vegetarian(1)/non-veg(2)? : '))

if(type == '1'):
    type='lunch,dinner'
elif(type == '2'):
    type='snacks'
elif(type == '3'):
    type='snacks,lunch,dinner'
elif(type == '4'):
    type='dessert'

if(diet == '1'):
    diet='vegetarian'
elif(diet == '2'):
    diet='non-veg'

html_texts = requests.get(link).text

soup = BeautifulSoup(html_texts, 'lxml')

#find the name of the recipe
name = soup.find('div', class_='recipe-title')
name = name.text.strip()

attr = soup.findAll('span', class_='attr-heading-value')

#find the serving
serving = attr[2].text

#find the time
times = attr[5]
times = times.text.strip()
times = re.split('\s', times)

print(times)

time = times[0]

if(times[len(times)-1] == 'hour' or times[len(times)-1] == 'hour(s)'):
    time = float(time) * 60

if len(times) > 4:
    time = float(times[0]) * 60 + float(times[4])

#find ingredients
ingredients = soup.findAll('div', class_='ingredient-description')
quantity = soup.findAll('div', class_='ingredient-quantity')

final_ingredients = ''

for i in range(len(quantity)):
    quantity[i] = quantity[i].text.strip()
    ingredients[i] = ingredients[i].text.strip()

    mymy = re.split('\s', ingredients[i])

    if(mymy[0] != 'water'):
        #print(f'{ingredients[i]} {quantity[i]}')

        a = re.sub('\s', '', quantity[i])

        if(a == ''): a = '5g'

        b = re.findall('[^0-9]', a)
        e = re.findall('[0-9/-/\/]', a)

        c = ''
        f = ''

        for d in b:
            if d == '-': d = 'g'
            c += d

        for g in e:
            f += g

        f = re.split('-', f)

        if '/' in c:
            c = c[1:len(c)]

        if c == 'cup':
            if f[0] == '1/2':
                f[0] = 0.5
            elif f[0] == '1/4':
                f[0] = 0.25
            elif f[0] == '3/4':
                f[0] = 0.75
            elif f[0] == '1/3':
                f[0] = 0.34
            elif f[0] == '11/2':
                f[0] = 1.5
            elif f[0] == '21/4':
                f[0] = 2.25
            elif f[0] == '13/4':
                f[0] = 1.75
            f[0] = float(f[0]) * 250
            c = 'g'
        elif c == 'grams':
            c = 'g'
        elif c == 'tablespoon':
            if f[0] == '1/4':
                f[0] = 0.25
            elif f[0] == '1/2':
                f[0] = 0.5
            elif f[0] == '3/4':
                f[0] = .75
            elif f[0] == '11/2':
                f[0] = 1.5
            elif f[0] == '41/2':
                f[0] = 4.5
            f[0] = float(f[0]) * 15
            c = 'g'
        elif c == 'teaspoon':
            if f[0] == '1/2':
                f[0] = 0.5
            elif f[0] == '1/4':
                f[0] = 0.25
            elif f[0] == '11/2':
                f[0] = 1.5
            elif f[0] == '3/4':
                f[0] = 0.75
            elif f[0] == '1/3':
                f[0] = 0.34
            elif f[0] == '41/2':
                f[0] = 4.5
            elif f[0] == '11/4':
                f[0] = 1.25
            f[0] = float(f[0]) * 4.2
            c = 'g'
        elif c == 'pinch':
            f[0] = float(f[0])
            c = 'g'
        elif c == 'smidgen':
            f[0] = float(f[0]) * 0.18
            c = 'g'
        elif c == 'liter':
            f[0] = float(f[0]) * 1000
            c = 'g'
        elif c == 'gallon':
            if f[0] == '3/4':
                f[0] = 0.75
            f[0] = f[0] * 1000
            c = 'g'
        elif c == '':
            c = 'g'

        final_ingredients += (str(f[0]) + c + ' ' + ingredients[i] + ', ')

final_ingredients = final_ingredients[:len(final_ingredients) - 2]

print(f'Name: {name}\nIngredients: {final_ingredients}\nTime(s): {time}\nDiet: {diet}\nServing: {serving}\nType: {type}\nLink: {link}\n')

sa = str(input('Save? y/n : '))

if(sa == 'y'):
    # save data in a excel file
    #wb = openpyxl.Workbook()
    wb = load_workbook('S_Recipe.xlsx')
    ws = wb.active

    data = (name, final_ingredients, time, diet, serving, type, link)

    ws.append(data)
    wb.save('S_Recipe.xlsx')


